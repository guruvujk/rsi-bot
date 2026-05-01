# paper_trade.py — Virtual trade simulator
# FIXES:
#   1. Removed broken sync_buy_to_api / sync_sell_to_api calls inside buy()/sell()
#      (those functions were defined BELOW the class — caused UnboundLocalError)
#   2. Removed duplicate _log_csv method (was defined twice — SyntaxError risk)
#   3. Removed dead run() loop (main.py handles scanning — no duplication)
#   4. Removed stale requests to localhost:8000 (no backend server running)
#   5. buy() now returns itype properly from get_instrument_type()
#   6. Clean separation: PaperTrader only trades + logs. main.py scans + alerts.

import csv, os
from datetime import datetime
from config import get_instrument_type, get_usd_inr_rate

LOG_FILE   = "logs/paper_trades.csv"
EXCEL_FILE = "logs/Trading_Journal.xlsx"


class PaperTrader:

    def __init__(self, capital: float):
        self.initial_capital = capital
        self.capital         = capital
        self.positions: dict = {}
        self.trades:    list = []
        self.open_pnl        = 0.0
        os.makedirs("logs", exist_ok=True)
        self._init_excel()

    # ── EXCEL INIT ───────────────────────────────────────────────────────────
    def _init_excel(self):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            if os.path.exists(EXCEL_FILE):
                try:
                    load_workbook(EXCEL_FILE, read_only=True).close()
                    return
                except Exception:
                    import shutil, time
                    backup = EXCEL_FILE.replace(".xlsx", f"_corrupted_{int(time.time())}.xlsx")
                    shutil.move(EXCEL_FILE, backup)
                    print(f"  ⚠️  Corrupt Excel backed up to: {backup}")
                    print(f"  🔄 Recreating fresh journal...")

            wb   = Workbook()
            thin = Side(style="thin", color="BDC3C7")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr  = Alignment(horizontal="center", vertical="center")

            def hdr(ws, row, col, val, bg="1E3A5F", fg="FFFFFF", sz=11):
                c            = ws.cell(row=row, column=col, value=val)
                c.font       = Font(name="Arial", size=sz, bold=True, color=fg)
                c.fill       = PatternFill("solid", fgColor=bg)
                c.alignment  = ctr
                c.border     = bdr

            # Sheet 1 — Trade Log
            ws1 = wb.active
            ws1.title = "Trade Log"
            ws1.sheet_view.showGridLines = False
            ws1.freeze_panes = "A3"
            ws1.merge_cells("A1:R1")
            c = ws1["A1"]
            c.value     = "📈  VIRTUAL TRADING JOURNAL — RSI BOT  (Auto Updated)"
            c.font      = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="1E3A5F")
            c.alignment = ctr
            ws1.row_dimensions[1].height = 28

            cols = [
                ("#",5),("Date",13),("Time",10),("Symbol",14),("Type",11),
                ("Action",9),("Buy Price",13),("Sell Price",13),("Qty",7),
                ("Investment ₹",15),("Stop Loss",13),("Target",12),
                ("P&L ₹",12),("Result",11),("RSI",8),
                ("MACD",10),("BB %",9),("Reason",22),
            ]
            for i, (h, w) in enumerate(cols, 1):
                hdr(ws1, 2, i, h)
                ws1.column_dimensions[get_column_letter(i)].width = w
            ws1.row_dimensions[2].height = 26

            # Sheet 2 — Daily Summary
            ws2 = wb.create_sheet("Daily Summary")
            ws2.sheet_view.showGridLines = False
            ws2.freeze_panes = "A3"
            ws2.merge_cells("A1:J1")
            c = ws2["A1"]
            c.value     = "📅  DAILY SUMMARY  (Auto Updated)"
            c.font      = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="2E86AB")
            c.alignment = ctr
            ws2.row_dimensions[1].height = 28
            dcols = [
                ("Date",13),("Trades",10),("Wins",8),("Losses",8),
                ("Win Rate %",13),("Daily P&L ₹",15),
                ("Cumulative P&L ₹",20),("Capital ₹",15),
                ("Best Trade ₹",15),("Notes",30),
            ]
            for i, (h, w) in enumerate(dcols, 1):
                hdr(ws2, 2, i, h, bg="2E86AB")
                ws2.column_dimensions[get_column_letter(i)].width = w
            ws2.row_dimensions[2].height = 26

            # Sheet 3 — By Instrument
            ws3 = wb.create_sheet("By Instrument")
            ws3.sheet_view.showGridLines = False
            ws3.merge_cells("A1:J1")
            c = ws3["A1"]
            c.value     = "🏆  PERFORMANCE BY INSTRUMENT TYPE  (Auto Updated)"
            c.font      = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="1ABC9C")
            c.alignment = ctr
            ws3.row_dimensions[1].height = 28
            icols = [
                ("Type",13),("Symbol",16),("Trades",10),("Wins",8),
                ("Losses",8),("Win Rate %",13),("Total P&L ₹",15),
                ("Avg P&L ₹",14),("Best ₹",13),("Worst ₹",13),
            ]
            for i, (h, w) in enumerate(icols, 1):
                hdr(ws3, 2, i, h, bg="1ABC9C")
                ws3.column_dimensions[get_column_letter(i)].width = w
            ws3.row_dimensions[2].height = 26

            # Sheet 4 — Stats
            ws4 = wb.create_sheet("Stats")
            ws4.sheet_view.showGridLines = False
            ws4.column_dimensions["A"].width = 24
            ws4.column_dimensions["B"].width = 20
            ws4.merge_cells("A1:B1")
            c = ws4["A1"]
            c.value     = "📊  LIVE STATS  (Auto Updated)"
            c.font      = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="8E44AD")
            c.alignment = ctr
            ws4.row_dimensions[1].height = 28
            slabels = [
                "Starting Capital ₹","Current Capital ₹","Total P&L ₹",
                "Open P&L ₹","Portfolio Value ₹","Return %",
                "Total Trades","Wins","Losses","Win Rate %",
                "Max Drawdown","Sharpe Ratio","Last Updated",
            ]
            for i, lbl in enumerate(slabels, 3):
                c = ws4.cell(row=i, column=1, value=lbl)
                c.font      = Font(name="Arial", size=11, bold=True, color="1E3A5F")
                c.fill      = PatternFill("solid", fgColor="D6EAF8" if i%2==0 else "EBF5FB")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border    = bdr
                ws4.row_dimensions[i].height = 22
                c2          = ws4.cell(row=i, column=2, value="—")
                c2.font     = Font(name="Arial", size=11)
                c2.fill     = PatternFill("solid", fgColor="D6EAF8" if i%2==0 else "EBF5FB")
                c2.alignment= ctr
                c2.border   = bdr

            ws1.sheet_properties.tabColor = "1E3A5F"
            ws2.sheet_properties.tabColor = "2E86AB"
            ws3.sheet_properties.tabColor = "1ABC9C"
            ws4.sheet_properties.tabColor = "8E44AD"
            wb.save(EXCEL_FILE)
            print(f"  ✅ Excel journal created: {EXCEL_FILE}")

        except ImportError:
            print("  ⚠️  openpyxl not installed — run: pip install openpyxl")
        except Exception as e:
            print(f"  ⚠️  Excel init error: {e}")

    # ── SAFE EXCEL SAVE (add this method to PaperTrader) ─────────────────────────
    def _safe_save(self, wb):
        """
        Atomic save with file lock.
        1. Acquires an exclusive lock on a .lock sidecar file
        2. Saves to a .tmp file
        3. Atomically replaces the real file
        This prevents corruption when two scan cycles overlap.
        """
        import fcntl, tempfile

        lock_path = EXCEL_FILE + ".lock"
        tmp_path  = EXCEL_FILE + ".tmp"

        with open(lock_path, "w") as lock_fh:
            try:
                fcntl.flock(lock_fh, fcntl.LOCK_EX)   # blocks until safe to proceed
                wb.save(tmp_path)
                os.replace(tmp_path, EXCEL_FILE)        # atomic on POSIX
            finally:
                fcntl.flock(lock_fh, fcntl.LOCK_UN)

    # ── LOG TRADE ROW TO EXCEL ────────────────────────────────────────────────
    def _log_excel_trade(self, row_data: dict):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb   = load_workbook(EXCEL_FILE)
            ws   = wb["Trade Log"]
            thin = Side(style="thin", color="BDC3C7")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr  = Alignment(horizontal="center", vertical="center")
            nr   = max(ws.max_row + 1, 3)
            bg   = "F2F7FC" if nr % 2 == 0 else "FFFFFF"
            bgf  = PatternFill("solid", fgColor=bg)

            vals = [
                nr - 2,
                row_data.get("date",""),
                row_data.get("time",""),
                row_data.get("symbol","").replace(".NS","").replace("=X","").replace("=F",""),
                row_data.get("itype",""),
                row_data.get("action",""),
                row_data.get("buy_price",""),
                row_data.get("sell_price",""),
                row_data.get("qty",""),
                row_data.get("investment",""),
                row_data.get("stop_loss",""),
                row_data.get("target",""),
                row_data.get("pnl",""),
                row_data.get("result",""),
                row_data.get("rsi",""),
                row_data.get("macd",""),
                row_data.get("bb_pct",""),
                row_data.get("reason",""),
            ]
            for col, val in enumerate(vals, 1):
                c           = ws.cell(row=nr, column=col, value=val)
                c.font      = Font(name="Arial", size=10)
                c.fill      = bgf
                c.alignment = ctr
                c.border    = bdr
                if col in [7, 8, 10, 11, 12, 13]:
                    c.number_format = "₹#,##0.00"
                if col == 6:   # Action
                    if val == "BUY":
                        c.font = Font(name="Arial", size=10, bold=True, color="1E8449")
                        c.fill = PatternFill("solid", fgColor="D5F5E3")
                    else:
                        c.font = Font(name="Arial", size=10, bold=True, color="922B21")
                        c.fill = PatternFill("solid", fgColor="FADBD8")
                if col == 13 and val != "":   # P&L
                    try:
                        v      = float(val)
                        c.font = Font(name="Arial", size=10, bold=True,
                                      color="1E8449" if v >= 0 else "922B21")
                    except: pass
                if col == 14 and val != "":   # Result
                    if "WIN" in str(val):
                        c.font = Font(name="Arial", size=10, bold=True, color="1E8449")
                        c.fill = PatternFill("solid", fgColor="D5F5E3")
                    elif "LOSS" in str(val):
                        c.font = Font(name="Arial", size=10, bold=True, color="922B21")
                        c.fill = PatternFill("solid", fgColor="FADBD8")
            ws.row_dimensions[nr].height = 22
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"  ⚠️  Excel trade log error: {e}")

    # ── UPDATE DAILY SUMMARY ──────────────────────────────────────────────────
    def _update_daily(self):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb    = load_workbook(EXCEL_FILE)
            ws    = wb["Daily Summary"]
            thin  = Side(style="thin", color="BDC3C7")
            bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr   = Alignment(horizontal="center", vertical="center")
            today = datetime.now().strftime("%Y-%m-%d")

            t_today = [t for t in self.trades if t.get("date","") == today]
            wins    = [t for t in t_today if t.get("pnl", 0) > 0]
            losses  = [t for t in t_today if t.get("pnl", 0) <= 0]
            dpnl    = sum(t.get("pnl", 0) for t in t_today)
            cpnl    = sum(t.get("pnl", 0) for t in self.trades)
            wr      = round(len(wins)/len(t_today)*100, 1) if t_today else 0
            best    = max((t.get("pnl", 0) for t in t_today), default=0)

            tr = None
            for row in ws.iter_rows(min_row=3):
                if row[0].value == today:
                    tr = row[0].row
                    break
            if tr is None:
                tr = max(ws.max_row + 1, 3)

            bg  = "F2F7FC" if tr % 2 == 0 else "FFFFFF"
            bgf = PatternFill("solid", fgColor=bg)
            vals = [today, len(t_today), len(wins), len(losses), wr,
                    round(dpnl,2), round(cpnl,2), round(self.capital,2),
                    round(best,2), ""]
            for col, val in enumerate(vals, 1):
                c           = ws.cell(row=tr, column=col, value=val)
                c.font      = Font(name="Arial", size=10)
                c.fill      = bgf
                c.alignment = ctr
                c.border    = bdr
                if col in [6, 7, 8, 9]:
                    c.number_format = "₹#,##0.00"
                if col == 5:
                    c.number_format = '0.0"%"'
                if col == 6:
                    c.font = Font(name="Arial", size=10, bold=True,
                                  color="1E8449" if dpnl >= 0 else "922B21")
            ws.row_dimensions[tr].height = 22
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"  ⚠️  Daily summary error: {e}")

    # ── UPDATE STATS SHEET ────────────────────────────────────────────────────
    def _update_stats(self):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            import statistics
            wb  = load_workbook(EXCEL_FILE)
            ws  = wb["Stats"]
            ctr = Alignment(horizontal="center", vertical="center")
            s   = self.stats()

            returns = [
                t.get("pnl", 0) / max(t.get("investment", 1), 1)
                for t in self.trades if t.get("action") == "SELL"
            ]
            sharpe = "—"
            if len(returns) > 1:
                avg    = statistics.mean(returns)
                std    = statistics.stdev(returns)
                sharpe = round(avg / std * (252**0.5), 2) if std > 0 else "—"

            vals = [
                f"₹{self.initial_capital:,.2f}",
                f"₹{s['capital']:,.2f}",
                f"₹{s['total_pnl']:,.2f}",
                f"₹{s['open_pnl']:,.2f}",
                f"₹{s['portfolio_val']:,.2f}",
                f"{s['return_pct']}%",
                s['total_trades'],
                s['wins'],
                s['losses'],
                f"{s['win_rate']}%",
                "—",
                sharpe,
                datetime.now().strftime("%d-%b-%Y %H:%M"),
            ]
            for i, val in enumerate(vals, 3):
                c           = ws.cell(row=i, column=2, value=val)
                c.alignment = ctr
                c.font      = Font(name="Arial", size=11)
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"  ⚠️  Stats update error: {e}")

    # ── UPDATE BY INSTRUMENT SHEET ────────────────────────────────────────────
    def _update_by_instrument(self):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb   = load_workbook(EXCEL_FILE)
            ws   = wb["By Instrument"]
            thin = Side(style="thin", color="BDC3C7")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr  = Alignment(horizontal="center", vertical="center")

            for row in ws.iter_rows(min_row=3):
                for cell in row:
                    cell.value = None

            sells = [t for t in self.trades if t.get("action") == "SELL"]
            by_sym: dict = {}
            for t in sells:
                sym = t.get("symbol", "")
                by_sym.setdefault(sym, []).append(t)

            TYPE_COLORS = {
                "STOCK":"D6EAF8","INDEX":"D5F5E3","ETF":"FEF9E7",
                "COMMODITY":"FDEBD0","FOREX":"F4ECF7",
                "CRYPTO":"E8F8F5","US_STOCK":"FDEDEC",
            }

            nr = 3
            for sym, trades in sorted(by_sym.items()):
                itype  = trades[0].get("itype", "STOCK")
                wins   = [t for t in trades if t.get("pnl", 0) > 0]
                losses = [t for t in trades if t.get("pnl", 0) <= 0]
                tpnl   = sum(t.get("pnl", 0) for t in trades)
                apnl   = tpnl / len(trades) if trades else 0
                bst    = max(t.get("pnl", 0) for t in trades)
                wst    = min(t.get("pnl", 0) for t in trades)
                wr     = round(len(wins)/len(trades)*100, 1) if trades else 0
                bg     = TYPE_COLORS.get(itype, "FFFFFF")
                bgf    = PatternFill("solid", fgColor=bg)
                vals   = [
                    itype,
                    sym.replace(".NS","").replace("=X","").replace("=F",""),
                    len(trades), len(wins), len(losses), wr,
                    round(tpnl,2), round(apnl,2), round(bst,2), round(wst,2),
                ]
                for col, val in enumerate(vals, 1):
                    c           = ws.cell(row=nr, column=col, value=val)
                    c.font      = Font(name="Arial", size=10)
                    c.fill      = bgf
                    c.alignment = ctr
                    c.border    = bdr
                    if col in [7, 8, 9, 10]:
                        c.number_format = "₹#,##0.00"
                    if col == 6:
                        c.number_format = '0.0"%"'
                ws.row_dimensions[nr].height = 22
                nr += 1
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"  ⚠️  By instrument error: {e}")

    # ── BUY ──────────────────────────────────────────────────────────────────
    def buy(self, symbol: str, price: float, qty: int, rsi: float,
            sl_pct: float = 0.02, tp_pct: float = 0.04,
            itype: str = None, indicators: dict = None) -> tuple:
        """
        Open a paper long position.
        Returns (True, message) on success, (False, reason) on failure.
        """
        cost  = qty * price
        if cost > self.capital:
            return False, f"Insufficient capital (need ₹{cost:,.0f})"

        # Auto-detect instrument type if not provided
        if not itype:
            itype = get_instrument_type(symbol)

        self.capital -= cost
        self.positions[symbol] = {
            "qty"      : qty,
            "buy_price": round(price, 4),
            "buy_time" : datetime.now().strftime("%H:%M"),
            "itype"    : itype,
            "stop_loss": round(price * (1 - sl_pct), 4),
        }

        ind = indicators or {}
        row = {
            "date"      : datetime.now().strftime("%Y-%m-%d"),
            "time"      : datetime.now().strftime("%H:%M"),
            "symbol"    : symbol,
            "itype"     : itype,
            "action"    : "BUY",
            "buy_price" : round(price, 4),
            "sell_price": "",
            "qty"       : qty,
            "investment": round(cost, 2),
            "stop_loss" : round(price * (1 - sl_pct), 4),
            "target"    : round(price * (1 + tp_pct), 4),
            "pnl"       : "",
            "result"    : "",
            "rsi"       : round(rsi, 2),
            "macd"      : round(ind.get("macd", 0), 4),
            "bb_pct"    : round(ind.get("bb_pct", 0), 2),
            "reason"    : "RSI BUY",
        }
        # Log to CSV + Excel only (no external API calls)
        self._log_csv(row)
        self._log_excel_trade(row)
        self._update_stats()
        return True, f"BUY {qty}x {symbol} @ {price:.4f}"

    # ── SELL ─────────────────────────────────────────────────────────────────
    def sell(self, symbol: str, price: float, rsi: float,
             reason: str, indicators: dict = None) -> tuple:
        """
        Close an open paper position.
        Returns (True, pnl) on success, (False, reason) on failure.
        """
        if symbol not in self.positions:
            return False, "No position"

        pos   = self.positions.pop(symbol)
        qty   = pos["qty"]
        itype = pos.get("itype", "STOCK")
        pnl   = round((price - pos["buy_price"]) * qty, 2)
        self.capital += qty * price

        ind   = indicators or {}
        trade = {
            "date"      : datetime.now().strftime("%Y-%m-%d"),
            "time"      : datetime.now().strftime("%H:%M"),
            "symbol"    : symbol,
            "itype"     : itype,
            "action"    : "SELL",
            "buy_price" : pos["buy_price"],
            "sell_price": round(price, 4),
            "qty"       : qty,
            "investment": round(pos["buy_price"] * qty, 2),
            "pnl"       : pnl,
            "result"    : "WIN 🟢" if pnl > 0 else "LOSS 🔴",
            "rsi"       : round(rsi, 2),
            "macd"      : round(ind.get("macd", 0), 4),
            "bb_pct"    : round(ind.get("bb_pct", 0), 2),
            "reason"    : reason,
        }
        self.trades.append(trade)
        self._log_csv(trade)
        self._log_excel_trade(trade)
        self._update_daily()
        self._update_stats()
        self._update_by_instrument()
        return True, pnl

    # ── OPEN P&L ─────────────────────────────────────────────────────────────
    def update_open_pnl(self, prices: dict) -> float:
        total = 0.0
        rate  = get_usd_inr_rate()
        for s, p in self.positions.items():
            if s not in prices:
                continue
            current = prices[s]
            bought  = p["buy_price"]
            qty     = p["qty"]
            if "-USD" in s:
                pnl = (current - bought) * qty * rate
            else:
                pnl = (current - bought) * qty
            total += pnl
        self.open_pnl = total
        return round(self.open_pnl, 2)

    # ── STOP-LOSS CHECKER ─────────────────────────────────────────────────────
    def check_and_exit_stops(self, prices: dict) -> list:
        """
        Check all open positions against STOP_LOSS_PCT.
        Calls self.sell() for any breach.
        Returns list of (symbol, pnl) for breached positions.
        """
        from config import STOP_LOSS_PCT, get_usd_inr_rate
        exited = []
        rate = get_usd_inr_rate()

        for symbol in list(self.positions.keys()):   # list() — safe to mutate during loop
            pos = self.positions.get(symbol)
            if not pos or symbol not in prices:
                continue

            ltp = prices[symbol]
            ltp_inr = ltp * rate if "-USD" in symbol else ltp
            buy_inr = pos["buy_price"] * rate if "-USD" in symbol else pos["buy_price"]
            stop_loss = buy_inr * (1 - STOP_LOSS_PCT)

            if ltp_inr <= stop_loss:
                loss_pct = (ltp_inr - buy_inr) / buy_inr
                print(f"  🔴 STOP-LOSS: {symbol} | "
      f"Buy ₹{buy_inr:.2f} → LTP ₹{ltp_inr:.2f} | "
      f"Stop ₹{stop_loss:.2f} | Loss {loss_pct*100:.2f}%")

                ok, pnl = self.sell(
                    symbol     = symbol,
                    price      = ltp,
                    rsi        = 50.0,        # RSI unknown at stop-loss time
                    reason     = f"STOP-LOSS ({loss_pct*100:.1f}%)",
                    indicators = {}
                )
                if ok:
                    exited.append((symbol, pnl))

        return exited

    # ── STATS ─────────────────────────────────────────────────────────────────
    def stats(self) -> dict:
        closed    = [t for t in self.trades if t.get("action") == "SELL"]
        wins      = [t for t in closed if t.get("pnl", 0) > 0]
        losses    = [t for t in closed if t.get("pnl", 0) <= 0]
        total_pnl = sum(t.get("pnl", 0) for t in closed)
        win_rate  = (len(wins) / len(closed) * 100) if closed else 0
        
        return {
            "total_trades" : len(closed),
            "wins"         : len(wins),
            "losses"       : len(losses),
            "win_rate"     : round(win_rate, 1),
            "total_pnl"    : round(total_pnl, 2),
            "open_pnl"     : round(self.open_pnl, 2),
            "capital"      : round(self.capital, 2),
            "portfolio_val": round(self.capital + self.open_pnl, 2),
            "return_pct"  : round((self.open_pnl / self.initial_capital) * 100, 2),
        }

    # ── CSV LOG ───────────────────────────────────────────────────────────────
    def _log_csv(self, row: dict):
        exists = os.path.exists(LOG_FILE)
        with open(LOG_FILE, "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=row.keys())
            if not exists:
                w.writeheader()
            w.writerow(row)