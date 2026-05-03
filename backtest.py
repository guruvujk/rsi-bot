# backtest.py — RSI Bot Backtesting Engine
# Tests full strategy: RSI + MACD + Volume filter + TSL
# Period: 2 years fetch, trade last 1 year | All instruments | Excel report

import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from config import (
    WATCHLIST, RSI_BUY, RSI_SELL, CAPITAL,
    get_instrument_type
)
from rsi_engine import fetch_ohlcv, compute_rsi, compute_macd, _safe_float

def get_adaptive_sl(itype: str) -> float:
    return {
        "FOREX"    : 0.008,
        "STOCK"    : 0.015,
        "INDEX"    : 0.010,
        "ETF"      : 0.015,
        "COMMODITY": 0.020,
        "CRYPTO"   : 0.040,
        "US_STOCK" : 0.020,
    }.get(itype, 0.02)

def get_adaptive_tp(itype: str) -> float:
    return {
        "FOREX"    : 0.006,
        "STOCK"    : 0.030,
        "INDEX"    : 0.020,
        "ETF"      : 0.020,
        "COMMODITY": 0.030,
        "CRYPTO"   : 0.050,
        "US_STOCK" : 0.030,
    }.get(itype, 0.03)


INITIAL_CAPITAL = CAPITAL
REPORT_FILE     = "logs/Backtest_Report.xlsx"
PERIOD          = "2y"   # fetch 2yr so EMA-200 is warmed up
INTERVAL        = "1d"


# ─────────────────────────────────────────────────────────────────────────────
def backtest_symbol(symbol: str) -> list:
    """Run backtest on a single symbol. Returns list of trade dicts."""
    try:
        df = yf.download(symbol, period=PERIOD, interval=INTERVAL,
                         progress=False, auto_adjust=True, threads=False)
        if df is None or df.empty or len(df) < 210:
            return []

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)

        df.dropna(inplace=True)
        close  = df['Close'].squeeze()
        volume = df['Volume'].squeeze() if 'Volume' in df.columns else None

        # ── Compute all indicators ONCE, outside the loop ────────────────────
        rsi                            = compute_rsi(close)
        macd_line, sig_line, histogram = compute_macd(close)
        vol_avg = volume.rolling(20).mean() if volume is not None else None
        ema200  = close.ewm(span=200, adjust=False).mean()

        itype  = get_instrument_type(symbol)
        sl_pct = get_adaptive_sl(itype)
        tp_pct = get_adaptive_tp(itype)

        trades   = []
        position = None
        capital  = INITIAL_CAPITAL

        # Start after EMA-200 is warmed up (bar 210 onwards)
        start_idx = max(210, len(df) - 252)

        for i in range(210, len(df)):
            price    = _safe_float(close.iloc[i])
            rsi_val  = _safe_float(rsi.iloc[i], 50.0)
            macd_val = _safe_float(macd_line.iloc[i])
            macd_s   = _safe_float(sig_line.iloc[i])
            date     = df.index[i]

            macd_bull        = macd_val > macd_s
            macd_hist_prev   = _safe_float(histogram.iloc[i - 1])
            macd_hist_curr   = _safe_float(histogram.iloc[i])
            macd_hist_rising = macd_hist_curr > macd_hist_prev

            # ── Volume filter ─────────────────────────────────────────────────
            vol_ok = True
            if vol_avg is not None:
                va     = _safe_float(vol_avg.iloc[i])
                vc     = _safe_float(volume.iloc[i])
                vol_ok = (vc > va) if va > 0 else True

            # ── Trend filter ──────────────────────────────────────────────────
            # Only use EMA-200: blocks bear markets, allows dip-buying in uptrends
            # price > EMA-50 is intentionally NOT required — RSI oversold means
            # the stock is dipping below short-term average, which is the entry point
            e200     = _safe_float(ema200.iloc[i])
            trend_ok = price > e200

            # ── EXIT ──────────────────────────────────────────────────────────
            if position:
                bp         = position['buy_price']
                highest    = position.get('highest_price', bp)
                tsl_active = position.get('tsl_active', False)
                chg_pct    = (price - bp) / bp

                if price > highest:
                    position['highest_price'] = price
                    highest = price

                if not tsl_active and chg_pct >= tp_pct * 0.4:
                    position['tsl_active'] = True
                    tsl_active = True

                if tsl_active:
                    profit    = highest - bp
                    tsl_price = bp + (profit * 0.5)
                    if tsl_price > position.get('stop_loss', 0):
                        position['stop_loss'] = tsl_price

                reason = None
                if price <= position['stop_loss']:
                    reason = "STOP LOSS"
                elif chg_pct >= tp_pct:
                    reason = "TARGET HIT"
                elif rsi_val > RSI_SELL and not macd_bull:
                    reason = "RSI SELL"

                if reason:
                    pnl     = (price - bp) * position['qty']
                    capital += position['qty'] * price
                    trades.append({
                        "symbol"    : symbol,
                        "itype"     : itype,
                        "buy_date"  : position['buy_date'].strftime("%d-%b-%Y"),
                        "sell_date" : date.strftime("%d-%b-%Y"),
                        "buy_price" : round(bp, 4),
                        "sell_price": round(price, 4),
                        "qty"       : position['qty'],
                        "pnl"       : round(pnl, 2),
                        "pnl_pct"   : round(chg_pct * 100, 2),
                        "reason"    : reason,
                        "tsl_used"  : tsl_active,
                        "result"    : "WIN" if pnl > 0 else "LOSS",
                    })
                    position = None

            # ── ENTRY ─────────────────────────────────────────────────────────
            # Buy the dip: RSI oversold + volume spike + MACD turning up
            # Only skip if price is below EMA-200 (confirmed bear market)
            elif rsi_val < RSI_BUY and vol_ok and macd_hist_rising and trend_ok:
                max_alloc = min(capital * 0.03, 5000)
                qty       = int(max_alloc / price)
                if qty <= 0 or qty * price > capital:
                    continue
                capital -= qty * price
                position = {
                    "buy_price"    : price,
                    "buy_date"     : date,
                    "qty"          : qty,
                    "stop_loss"    : round(price * (1 - sl_pct), 4),
                    "highest_price": price,
                    "tsl_active"   : False,
                }

        return trades

    except Exception as e:
        print(f"  [BT] {symbol}: {e}")
        return []


# ─────────────────────────────────────────────────────────────────────────────
def generate_excel_report(all_trades: list):
    """Generate Excel report with summary + charts."""
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    os.makedirs("logs", exist_ok=True)
    wb   = Workbook()
    thin = Side(style="thin", color="BDC3C7")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")

    def hdr(ws, row, col, val, bg="1E3A5F", fg="FFFFFF", sz=11):
        c           = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", size=sz, bold=True, color=fg)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = ctr
        c.border    = bdr

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1       = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:H1")
    c           = ws1["A1"]
    c.value     = "📊  RSI BOT — 1 YEAR BACKTEST REPORT"
    c.font      = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1E3A5F")
    c.alignment = ctr
    ws1.row_dimensions[1].height = 36

    if not all_trades:
        ws1["A2"] = "No trades generated in backtest period."
        wb.save(REPORT_FILE)
        print("  ⚠️  No trades found.")
        return

    df_t      = pd.DataFrame(all_trades)
    wins      = df_t[df_t['result'] == 'WIN']
    losses    = df_t[df_t['result'] == 'LOSS']
    total_pnl = df_t['pnl'].sum()
    win_rate  = len(wins) / len(df_t) * 100 if len(df_t) > 0 else 0
    avg_win   = wins['pnl'].mean()   if len(wins)   > 0 else 0
    avg_loss  = losses['pnl'].mean() if len(losses) > 0 else 0
    tsl_saves = df_t[df_t['tsl_used'] == True]['result'].value_counts().get('WIN', 0)

    stats = [
        ("Total Trades",       len(df_t)),
        ("Wins",               len(wins)),
        ("Losses",             len(losses)),
        ("Win Rate %",         f"{win_rate:.1f}%"),
        ("Total P&L ₹",        f"₹{total_pnl:,.2f}"),
        ("Avg Win ₹",          f"₹{avg_win:,.2f}"),
        ("Avg Loss ₹",         f"₹{avg_loss:,.2f}"),
        ("TSL Saves (wins)",   tsl_saves),
        ("Return %",           f"{(total_pnl / INITIAL_CAPITAL * 100):.2f}%"),
        ("Period",             "1 Year (2yr fetch, EMA-200 warmed up)"),
        ("Instruments tested", df_t['symbol'].nunique()),
        ("Strategy",           "RSI + MACD + Volume + EMA-200 + TSL"),
    ]

    ws1.row_dimensions[2].height = 10
    for i, (label, val) in enumerate(stats, 3):
        bg = "EBF5FB" if i % 2 == 0 else "FFFFFF"
        for col, v in [(1, label), (2, val)]:
            c           = ws1.cell(row=i, column=col, value=v)
            c.font      = Font(name="Arial", size=11,
                               bold=(col == 1),
                               color="1E3A5F" if col == 1 else "2C3E50")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center"
            )
            c.border    = bdr
        ws1.row_dimensions[i].height = 22
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22

    # ── Sheet 2: All Trades ───────────────────────────────────────────────────
    ws2       = wb.create_sheet("All Trades")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:K1")
    c           = ws2["A1"]
    c.value     = "📋  ALL BACKTEST TRADES"
    c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="2E86AB")
    c.alignment = ctr
    ws2.row_dimensions[1].height = 28

    cols = [
        ("Symbol", 9), ("Type", 10), ("Buy Date", 13), ("Sell Date", 13),
        ("Buy ₹", 12), ("Sell ₹", 12), ("Qty", 7),
        ("P&L ₹", 12), ("P&L %", 9), ("TSL", 7), ("Result", 10),
    ]
    for i, (h, w) in enumerate(cols, 1):
        hdr(ws2, 2, i, h, bg="2E86AB")
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 24

    for nr, t in enumerate(all_trades, 3):
        bg  = "F2F7FC" if nr % 2 == 0 else "FFFFFF"
        bgf = PatternFill("solid", fgColor=bg)
        row = [
            t['symbol'].replace('.NS', '').replace('=F', '').replace('-USD', '/USD'),
            t['itype'], t['buy_date'], t['sell_date'],
            t['buy_price'], t['sell_price'], t['qty'],
            t['pnl'], t['pnl_pct'],
            "✅" if t['tsl_used'] else "—",
            t['result'],
        ]
        for col, val in enumerate(row, 1):
            c           = ws2.cell(row=nr, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = bgf
            c.alignment = ctr
            c.border    = bdr
            if col == 8:
                c.number_format = "₹#,##0.00"
                c.font = Font(name="Arial", size=10, bold=True,
                              color="1E8449" if t['pnl'] >= 0 else "922B21")
            if col == 11:
                if val == "WIN":
                    c.font = Font(name="Arial", size=10, bold=True, color="1E8449")
                    c.fill = PatternFill("solid", fgColor="D5F5E3")
                else:
                    c.font = Font(name="Arial", size=10, bold=True, color="922B21")
                    c.fill = PatternFill("solid", fgColor="FADBD8")
        ws2.row_dimensions[nr].height = 20

    # ── Sheet 3: By Instrument ────────────────────────────────────────────────
    ws3       = wb.create_sheet("By Instrument")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:H1")
    c           = ws3["A1"]
    c.value     = "🏆  PERFORMANCE BY INSTRUMENT"
    c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1ABC9C")
    c.alignment = ctr
    ws3.row_dimensions[1].height = 28

    icols = [
        ("Symbol", 14), ("Type", 11), ("Trades", 9), ("Wins", 7),
        ("Losses", 8), ("Win Rate %", 12), ("Total P&L ₹", 14), ("Avg P&L ₹", 13),
    ]
    for i, (h, w) in enumerate(icols, 1):
        hdr(ws3, 2, i, h, bg="1ABC9C")
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[2].height = 24

    by_sym = df_t.groupby('symbol')
    nr     = 3
    for sym, grp in sorted(by_sym, key=lambda x: x[1]['pnl'].sum(), reverse=True):
        w   = len(grp[grp['result'] == 'WIN'])
        l   = len(grp[grp['result'] == 'LOSS'])
        wr  = round(w / len(grp) * 100, 1)
        tp  = round(grp['pnl'].sum(), 2)
        ap  = round(grp['pnl'].mean(), 2)
        bg  = "F2F7FC" if nr % 2 == 0 else "FFFFFF"
        bgf = PatternFill("solid", fgColor=bg)
        row = [
            sym.replace('.NS', '').replace('=F', '').replace('-USD', '/USD'),
            grp['itype'].iloc[0], len(grp), w, l, wr, tp, ap
        ]
        for col, val in enumerate(row, 1):
            c           = ws3.cell(row=nr, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = bgf
            c.alignment = ctr
            c.border    = bdr
            if col in [7, 8]:
                c.number_format = "₹#,##0.00"
                c.font = Font(name="Arial", size=10, bold=True,
                              color="1E8449" if val >= 0 else "922B21")
        ws3.row_dimensions[nr].height = 20
        nr += 1

    # ── Sheet 4: P&L Chart ────────────────────────────────────────────────────
    ws4       = wb.create_sheet("P&L Chart")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "Trade #"
    ws4["B1"] = "Cumulative P&L"
    cum_pnl   = 0
    for i, t in enumerate(all_trades, 2):
        cum_pnl += t['pnl']
        ws4.cell(row=i, column=1, value=i - 1)
        ws4.cell(row=i, column=2, value=round(cum_pnl, 2))

    chart              = LineChart()
    chart.title        = "Cumulative P&L"
    chart.style        = 10
    chart.y_axis.title = "P&L (₹)"
    chart.x_axis.title = "Trade #"
    data = Reference(ws4, min_col=2, min_row=1, max_row=len(all_trades) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.width  = 20
    chart.height = 12
    ws4.add_chart(chart, "D2")

    ws1.sheet_properties.tabColor = "1E3A5F"
    ws2.sheet_properties.tabColor = "2E86AB"
    ws3.sheet_properties.tabColor = "1ABC9C"
    ws4.sheet_properties.tabColor = "E74C3C"

    wb.save(REPORT_FILE)
    print(f"  ✅ Report saved: {REPORT_FILE}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  RSI BOT — BACKTESTING ENGINE")
    print(f"  Fetch: 2yr | Trade window: last 1yr | EMA-200 warmed up")
    print(f"  Instruments: {len(WATCHLIST)}")
    print("=" * 60)

    all_trades = []
    for i, symbol in enumerate(WATCHLIST, 1):
        print(f"  [{i:>3}/{len(WATCHLIST)}] {symbol:<20}", end=" ")
        trades = backtest_symbol(symbol)
        all_trades.extend(trades)
        print(f"→ {len(trades)} trades")

    print(f"\n  Total trades: {len(all_trades)}")
    print(f"  Generating Excel report...")
    generate_excel_report(all_trades)

    if all_trades:
        df   = pd.DataFrame(all_trades)
        wins = df[df['result'] == 'WIN']
        print(f"\n  Win Rate : {len(wins) / len(df) * 100:.1f}%")
        print(f"  Total P&L: ₹{df['pnl'].sum():,.2f}")
        print(f"  Return   : {df['pnl'].sum() / CAPITAL * 100:.2f}%")
    print("=" * 60)