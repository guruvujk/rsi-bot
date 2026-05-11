from flask import Flask, render_template_string, request, jsonify
from flask_socketio import SocketIO
import threading, time

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")
from auto_trade_routes import auto_trade_bp
app.register_blueprint(auto_trade_bp)

# Load from PostgreSQL on startup
try:
    from db_state import load_state as _db_load, init_db as _init_db, load_trades as _db_load_trades
    _init_db()
    _saved = _db_load()
    if _saved:
        print(f"[Dashboard] Loaded from DB: capital={_saved.get('capital')}, positions={len(_saved.get('positions',{}))}, trades={len(_saved.get('trades',[]))}")
except Exception as _e:
    print(f"[Dashboard] DB load error: {_e}")
    _saved = None

bot_state = {
    "capital": _saved.get("capital", 100000) if _saved else 100000,
    "positions": _saved.get("positions", {}) if _saved else {},
    "trades": [],
    "pnl": _saved.get("pnl", 0.0) if _saved else 0.0,
    "open_pnl": _saved.get("open_pnl", 0.0) if _saved else 0.0,
    "realised_pnl": _saved.get("realised_pnl", 0.0) if _saved else 0.0,
    "total_trades": _saved.get("total_trades", 0) if _saved else 0,
    "wins": _saved.get("wins", 0) if _saved else 0,
    "losses": _saved.get("losses", 0) if _saved else 0,
    "win_rate": _saved.get("win_rate", 0) if _saved else 0,
    "return_pct": _saved.get("return_pct", 0) if _saved else 0,
    "watchlist": _saved.get("watchlist", {}) if _saved else {},
    "paper_mode": True,
}

@app.route('/api/trade', methods=['POST'])
def api_trade():
    try:
        data = request.get_json(silent=True) or {}
        return jsonify({"status": "ok", "received": data}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/position/update', methods=['POST'])
def api_position_update():
    try:
        data = request.get_json(silent=True) or {}
        symbol = data.get("symbol", "")
        price = data.get("current_price", 0)
        if symbol and symbol in bot_state.get("positions", {}):
            bot_state["positions"][symbol]["current_price"] = float(price)
        return jsonify({"status": "ok"}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/health')
def api_health():
    return jsonify({"status": "running", "paper_mode": True}), 200

@app.route('/api/position/remove', methods=['POST'])
def remove_position():
    try:
        data = request.get_json(silent=True) or {}
        symbol = data.get("symbol", "")
        if not symbol:
            return jsonify({"error": "No symbol provided"}), 400
        if symbol in bot_state.get("positions", {}):
            del bot_state["positions"][symbol]
        try:
            from db_state import save_state as db_save
            db_save(dict(bot_state))
        except Exception as e:
            print(f"[Remove] DB error: {e}")
        return jsonify({"status": "ok", "removed": symbol})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

#  Main dashboard page 
DASHBOARD_HTML = """
<!DOCTYPE html>
<html>
<head>
    <title>RSI Bot -- Paper Trade Dashboard</title>
    <script src="/static/socket.io.min.js"></script>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Segoe UI', sans-serif; background: #f5f7fa; color: #1a1a2e; }
        .header { background: #fff; padding: 14px 28px; border-bottom: 1px solid #e2e8f0;
                  display: flex; align-items: center; justify-content: space-between;
                  box-shadow: 0 1px 4px rgba(0,0,0,0.06); }
        .header h1 { font-size: 18px; font-weight: 600; color: #2563eb; }
        .paper-badge { background: #eff6ff; color: #2563eb; font-size: 11px; font-weight: 600;
                       padding: 3px 10px; border-radius: 20px; border: 1px solid #bfdbfe; }
        .live-dot { width: 9px; height: 9px; border-radius: 50%; background: #16a34a;
                    animation: pulse 1.5s infinite; display: inline-block; margin-right: 6px; }
        @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:0.4} }
        .stats { display: grid; grid-template-columns: repeat(4,1fr); gap: 14px; padding: 20px 28px; }
        .stats2 { display: grid; grid-template-columns: repeat(4,1fr); gap: 14px; padding: 0 28px 20px; }
        .stat { background: #fff; border: 1px solid #e2e8f0; border-radius: 10px;
                padding: 16px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }
        .stat2 { background: #fff; border: 1px solid #e2e8f0; border-radius: 10px;
                 padding: 12px 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.05);
                 display: flex; flex-direction: column; align-items: center; }
        .label { font-size: 11px; color: #64748b; text-transform: uppercase;
                 letter-spacing: 0.8px; margin-bottom: 6px; }
        .value { font-size: 22px; font-weight: 600; }
        .stat2 .value { font-size: 20px; }
        .note { margin-top: 8px; font-size: 11px; color: #64748b; text-align: center; }
        .green { color: #16a34a; }
        .red { color: #dc2626; }
        .blue { color: #2563eb; }
        .gray { color: #475569; }
        .grid2 { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; padding: 0 28px 20px; }
        .grid1 { display: grid; grid-template-columns: 1fr; gap: 14px; padding: 0 28px 20px; }
        .box { background: #fff; border: 1px solid #e2e8f0; border-radius: 10px;
               overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }
        .box-title { padding: 11px 16px; font-size: 12px; font-weight: 600; color: #64748b;
                     border-bottom: 1px solid #f1f5f9; text-transform: uppercase;
                     letter-spacing: 0.5px; background: #f8fafc;
                     display: flex; align-items: center; justify-content: space-between; }
        table { width: 100%; border-collapse: collapse; font-size: 13px; }
        th { padding: 9px 16px; text-align: left; font-size: 11px; color: #94a3b8;
             border-bottom: 1px solid #f1f5f9; font-weight: 600; background: #f8fafc; }
        td { padding: 9px 16px; border-bottom: 1px solid #f8fafc; color: #334155; }
        tr:last-child td { border-bottom: none; }
        tr:hover td { background: #f8fafc; }
        .badge { display: inline-block; padding: 2px 10px; border-radius: 20px;
                 font-size: 11px; font-weight: 600; }
        .badge-buy { background: #dcfce7; color: #16a34a; }
        .badge-sell { background: #fee2e2; color: #dc2626; }
        .badge-hold { background: #f1f5f9; color: #64748b; }
        .rsi-bar-wrap { width: 70px; height: 7px; background: #e2e8f0; border-radius: 4px;
                        display: inline-block; vertical-align: middle; }
        .rsi-bar { height: 100%; border-radius: 4px; }
        .win-bar-wrap { width: 100%; height: 8px; background: #fee2e2; border-radius: 4px; margin-top: 6px; }
        .win-bar { height: 100%; border-radius: 4px; background: #16a34a; }
        .empty-msg { color: #94a3b8; text-align: center; padding: 24px; font-size: 13px; }
        button { cursor: pointer; }
    </style>
</head>
<body>

<div class="header">
    <div style="display:flex;align-items:center;gap:12px;">
        <h1> RSI Algo Bot</h1>
        <span class="paper-badge"> Paper Trade</span>
    </div>
    <div style="font-size:13px;color:#64748b;">
        <span class="live-dot"></span>Live &nbsp;|&nbsp;
        <span id="clock">--:--:--</span>
    </div>
</div>

<div class="stats">
    <div class="stat"><div class="label">Available Capital</div><div class="value blue" id="capital">...</div></div>
    <div class="stat"><div class="label">Portfolio Value</div><div class="value" id="portfolio_val">...</div></div>
    <div class="stat"><div class="label">Realised P&L</div><div class="value green" id="pnl">Rs.0</div></div>
    <div class="stat"><div class="label">Unrealised P&L</div><div class="value gray" id="open_pnl">Rs.0</div></div>
</div>

<div class="stats2">
    <div class="stat2"><div class="label">Closed Trades</div><div class="value gray" id="total_trades">0</div><div class="note">SELL trades only</div></div>
    <div class="stat2"><div class="label">Win Rate</div><div class="value green" id="win_rate">0%</div><div class="win-bar-wrap"><div class="win-bar" id="win-bar" style="width:0%"></div></div></div>
    <div class="stat2"><div class="label">Wins / Losses</div><div class="value" id="wins_losses">0 / 0</div></div>
    <div class="stat2"><div class="label">Return</div><div class="value gray" id="return_pct">0%</div><div class="note">Includes open position cost</div></div>
</div>

<div class="grid2">
    <!-- Watchlist Box -->
    <div class="box">
        <div class="box-title">Watchlist -- RSI Scanner <span id="wl-count" style="font-size:11px;color:#94a3b8;"></span></div>
        <table>
            <thead><tr><th>Symbol</th><th>Price</th><th>RSI</th><th>Signal</th></tr></thead>
            <tbody id="watchlist-body"><tr><td colspan="4" class="empty-msg">Waiting for scan...</td></tr></tbody>
        </table>
    </div>

    <!-- Open Positions Box - REAL POSITIONS ONLY -->
    <div class="box">
        <div class="box-title">Real Portfolio -- Live Positions <span id="pos-count" style="font-size:11px;color:#94a3b8;"></span></div>
        <table>
            <thead>
                <tr>
                    <th>Symbol</th><th>Qty</th><th>Buy @</th><th>LTP</th><th>P&L</th><th>Broker</th><th></th>
                </tr>
            </thead>
            <tbody id="positions-body">
                <tr><td colspan="7" class="empty-msg">No real positions found. Connect broker (Kite/Groww/Upstox)</td></tr>
            </tbody>
        </table>
    </div>
</div>

<!-- Trade Log -->
<div class="grid1">
    <div class="box">
        <div class="box-title">Trade Log (Last 20)</div>
        <table>
            <thead><tr><th>Date</th><th>Time</th><th>Symbol</th><th>Action</th><th>Price</th><th>Qty</th><th>RSI</th><th>P&L</th><th>Reason</th></tr></thead>
            <tbody id="trade-log"><tr><td colspan="9" class="empty-msg">No trades yet</td></tr></tbody>
        </table>
    </div>
</div>

<!-- Add Manual Position (Real Broker Only) -->
<div class="grid1">
    <div class="box">
        <div class="box-title" style="background:#f0fdf4;border-bottom:1px solid #bbf7d0;">
            <span style="color:#16a34a;"> Add Real Position (Kite / Groww / Upstox / Zerodha)</span>
        </div>
        <div style="padding:14px 16px;display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;">
            <div><label style="font-size:11px;">Symbol</label><input id="m-symbol" placeholder="SUNPHARMA.NS" style="padding:6px 10px;border:1px solid #e2e8f0;border-radius:6px;width:120px;"></div>
            <div><label style="font-size:11px;">Qty</label><input id="m-qty" type="number" placeholder="2" style="padding:6px 10px;border:1px solid #e2e8f0;border-radius:6px;width:80px;"></div>
            <div><label style="font-size:11px;">Buy Price</label><input id="m-price" type="number" placeholder="1852.30" style="padding:6px 10px;border:1px solid #e2e8f0;border-radius:6px;width:100px;"></div>
            <div><label style="font-size:11px;">Broker</label><select id="m-broker" style="padding:6px 10px;border:1px solid #e2e8f0;border-radius:6px;"><option>Kite</option><option>Groww</option><option>Upstox</option><option>Zerodha</option></select></div>
            <div><label style="font-size:11px;">Type</label><select id="m-itype" style="padding:6px 10px;border:1px solid #e2e8f0;border-radius:6px;"><option value="STOCK">STOCK</option><option value="US_STOCK">US_STOCK</option></select></div>
            <button onclick="addRealPosition()" style="padding:7px 18px;background:#16a34a;color:#fff;border:none;border-radius:6px;font-weight:600;"> Add Real Position</button>
            <span id="m-msg" style="font-size:12px;color:#64748b;"></span>
        </div>
    </div>
</div>

<!-- Upstox Real Portfolio Section -->
<div class="grid1">
    <div class="box">
        <div class="box-title" style="background:#eff6ff;border-bottom:1px solid #bfdbfe;">
            <span style="color:#2563eb;"> All Real Portfolios (Kite / Groww / Upstox / Zerodha)</span>
            <div>
                <span id="upstox-token-status" style="font-size:11px;padding:2px 10px;border-radius:20px;background:#dcfce7;color:#16a34a;">Ready</span>
                <button onclick="upstoxLogin()" style="font-size:11px;padding:4px 12px;border-radius:6px;border:1px solid #bfdbfe;background:#fff;color:#2563eb;"> Upstox Login</button>
                <button onclick="syncAllBrokers()" style="font-size:11px;padding:4px 12px;border-radius:6px;border:none;background:#2563eb;color:#fff;"> Sync All</button>
            </div>
        </div>
        <table>
            <thead><tr><th>Symbol</th><th>Type</th><th>Qty</th><th>Buy @</th><th>LTP</th><th>P&L</th><th>Broker</th><th>Synced</th><th></th></tr></thead>
            <tbody id="upstox-body"><tr><td colspan="9" class="empty-msg">No real positions. Add via form above or sync broker.</td></tr></tbody>
        </table>
        <div id="upstox-sync-msg" style="padding:8px 16px;font-size:12px;color:#64748b;display:none;"></div>
    </div>
</div>

<script>
// Add real position (paper_mode = false)
/*
function addRealPosition() {
    const symbol = document.getElementById('m-symbol').value.trim().toUpperCase();
    const qty = document.getElementById('m-qty').value;
    const buy_price = document.getElementById('m-price').value;
    const broker = document.getElementById('m-broker').value;
    const itype = document.getElementById('m-itype').value;
    const msg = document.getElementById('m-msg');
    
    if (!symbol || !qty || !buy_price) {
        msg.textContent = 'X Fill all fields';
        msg.style.color = '#dc2626';
        return;
    }
    
    msg.textContent = '... Adding real position...';
    fetch('/api/auto/manual/add', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({symbol, qty: parseInt(qty), buy_price: parseFloat(buy_price), broker, itype})
    }).then(r => r.json()).then(d => {
        if (d.status === 'added') {
            msg.textContent = 'OK Real position added! P&L will update live.';
            msg.style.color = '#16a34a';
            document.getElementById('m-symbol').value = '';
            document.getElementById('m-qty').value = '';
            document.getElementById('m-price').value = '';
            loadRealPositions();
        } else {
            msg.textContent = 'X ' + (d.error || 'Failed');
            msg.style.color = '#dc2626';
        }
    });
}
*/

function upstoxLogin() {
    fetch('/api/auto/upstox/login').then(r => r.json()).then(d => { window.open(d.login_url, '_blank'); });
}

function syncAllBrokers() {
    const msg = document.getElementById('upstox-sync-msg');
    msg.style.display = 'block';
    msg.textContent = '... Syncing Upstox...';
    fetch('/api/auto/upstox/sync', {method:'POST'})
        .then(r => r.json())
        .then(d => { 
            msg.textContent = 'OK Synced at ' + new Date().toLocaleTimeString('en-IN'); 
            loadRealPositions();
        })
        .catch(() => { msg.textContent = 'X Sync failed'; });
}

function loadRealPositions() {
    fetch('/api/auto/upstox/positions')
        .then(function(r){ return r.json(); })
        .then(function(d){
            var tbody = document.getElementById('positions-body');
            var positions = d.positions || [];
            if (positions.length === 0) {
                tbody.innerHTML = '<tr><td colspan="7" class="empty-msg">No real positions. Add via form above.</td></tr>';
                return;
            }
            var html = '';
            for (var i = 0; i < positions.length; i++) {
                var p = positions[i];
                var buyFmt = p.itype === 'US_STOCK' ? '$' + Number(p.buy_price).toFixed(2) : 'Rs.' + Number(p.buy_price).toLocaleString('en-IN');
                var ltpFmt = p.itype === 'US_STOCK' ? '$' + Number(p.ltp).toFixed(2) : 'Rs.' + Number(p.ltp).toLocaleString('en-IN');
                var pnlFmt = (p.pnl >= 0 ? '+' : '-') + 'Rs.' + Math.abs(p.pnl).toFixed(2);
                var pnlColor = p.pnl > 0 ? '#16a34a' : p.pnl < 0 ? '#dc2626' : '#64748b';
                var brokerColor = p.broker === 'Upstox' ? '#2563eb' : p.broker === 'Kite' ? '#16a34a' : p.broker === 'Groww' ? '#854d0e' : '#64748b';
                var brokerBg = p.broker === 'Upstox' ? '#eff6ff' : p.broker === 'Kite' ? '#f0fdf4' : p.broker === 'Groww' ? '#fef9c3' : '#f1f5f9';
                html += '<tr>' +
                    '<td style="font-weight:600">' + p.symbol + '</td>' +
                    '<td>' + p.qty + '</td>' +
                    '<td>' + buyFmt + '</td>' +
                    '<td>' + ltpFmt + '</td>' +
                    '<td style="font-weight:600;color:' + pnlColor + '">' + pnlFmt + '</td>' +
                    '<td><span class="badge" style="background:' + brokerBg + ';color:' + brokerColor + '">' + (p.broker || 'Unknown') + '</span></td>' +
                    '<td><button data-sym="' + p.symbol + '" onclick="removeRealPosition(this.dataset.sym)" style="font-size:10px;padding:2px 8px;border:1px solid #fca5a5;background:#fff;color:#dc2626;border-radius:4px;">X</button></td>' +
                    '</tr>';
            }
            tbody.innerHTML = html;
        });
}

function removeRealPosition(symbol) {
    if (!confirm('Remove ' + symbol + ' from real portfolio?')) return;
    fetch('/api/position/remove', {method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({symbol:symbol})})
        .then(r => r.json()).then(d => { if(d.status==='ok'){alert(symbol+' removed');loadRealPositions();} else{alert('Error');} });
}

// Load on page open
setTimeout(loadRealPositions, 2000);
setInterval(loadRealPositions, 30000);
var socket = io();
function fmt(n) { return 'Rs.' + Number(n).toLocaleString('en-IN', {maximumFractionDigits:2}); }
function colorVal(el, val) { el.className = 'value ' + (val > 0 ? 'green' : val < 0 ? 'red' : 'gray'); }

setInterval(() => { document.getElementById('clock').textContent = new Date().toLocaleTimeString('en-IN'); }, 1000);

socket.on('state_update', (d) => {
    document.getElementById('capital').textContent = fmt(d.capital || 0);
    document.getElementById('portfolio_val').textContent = fmt(d.portfolio_val || d.capital || 0);
    const pnlEl = document.getElementById('pnl'); pnlEl.textContent = fmt(d.pnl || 0); colorVal(pnlEl, d.pnl || 0);
    const opEl = document.getElementById('open_pnl'); opEl.textContent = fmt(d.open_pnl || 0); colorVal(opEl, d.open_pnl || 0);
    const retEl = document.getElementById('return_pct'); retEl.textContent = (d.return_pct || 0) + '%'; colorVal(retEl, d.return_pct || 0);
    document.getElementById('total_trades').textContent = d.total_trades || 0;
    document.getElementById('win_rate').textContent = (d.win_rate || 0) + '%';
    document.getElementById('wins_losses').textContent = (d.wins||0) + ' / ' + (d.losses||0);
    document.getElementById('win-bar').style.width = Math.min(d.win_rate||0, 100) + '%';
    
    // Watchlist
    const wl = d.watchlist || {};
    document.getElementById('wl-count').textContent = Object.keys(wl).length + ' stocks';
    let whtml = '';
    for (const [sym, v] of Object.entries(wl)) {
        const rsiColor = v.rsi < 30 ? '#16a34a' : v.rsi > 70 ? '#dc2626' : '#64748b';
        const rsiPct = Math.min(v.rsi, 100);
        const bClass = v.signal === 'BUY' ? 'badge-buy' : v.signal === 'SELL' ? 'badge-sell' : 'badge-hold';
        const isUsd = sym.includes('-USD') || sym.includes('/USD');
        const priceStr = v.price ? (isUsd ? '$' + Number(v.price).toLocaleString('en-US', {maximumFractionDigits:2}) : fmt(v.price)) : '--';
        whtml += `<tr><td style="font-weight:500;">${sym.replace('.NS','')}</td><td>${priceStr}</td><td><span style="color:${rsiColor};font-weight:600;">${Number(v.rsi).toFixed(1)}</span><div class="rsi-bar-wrap"><div class="rsi-bar" style="width:${rsiPct}%;background:${rsiColor};"></div></div></td><td><span class="badge ${bClass}">${v.signal}</span></td></tr>`;
    }
    document.getElementById('watchlist-body').innerHTML = whtml || '<tr><td colspan="4" class="empty-msg">Waiting for scan...</td></tr>';
    
    // Trade Log
    const trades = [...(d.trades || [])].reverse().slice(0, 20);
    let thtml = '';
    trades.forEach(t => {
        const pnlStr = t.pnl != null ? `<span style="color:${t.pnl>=0?'#16a34a':'#dc2626'};">${fmt(t.pnl)}</span>` : '--';
        const bClass = t.action === 'BUY' ? 'badge-buy' : 'badge-sell';
        thtml += `<tr><td style="color:#94a3b8;">${t.date || ''}</td><td style="color:#94a3b8;">${t.time || ''}</td><td>${String(t.symbol||'').replace('.NS','')}</td><td><span class="badge ${bClass}">${(t.action||'').toUpperCase()}</span></td><td>${fmt(t.price || t.buy_price || 0)}</td><td>${t.qty || 0}</td><td>${Number(t.rsi||0).toFixed(1)}</td><td>${pnlStr}</td><td style="color:#94a3b8;">${t.reason || ''}</td></tr>`;
    });
    document.getElementById('trade-log').innerHTML = thtml || '<tr><td colspan="9" class="empty-msg">No trades yet</td></tr>';
});
</script>
</body>
</html>
"""

@app.route("/api/positions")
def api_positions():
    from dashboard import bot_state as s
    raw = s.get("positions", {})
    pos_list = []
    if isinstance(raw, dict):
        for sym, p in raw.items():
            pos_list.append({"symbol": sym, "qty": p.get("qty",0), "buy_price": p.get("buy_price",0), "ltp": p.get("current_price", p.get("buy_price",0)), "pnl": p.get("pnl",0), "stop_loss": p.get("stop_loss",0)})
    return jsonify({"positions": pos_list, "count": len(pos_list)})

@app.route("/portfolio")
def portfolio():
    try:
        from auto_trade_engine import get_portfolio_summary
        data = get_portfolio_summary()
        capital = bot_state.get("capital", 0)
        return jsonify({"capital": capital, "portfolio_value": capital + data.get("total_pnl", 0), "unrealised_pnl": data.get("total_pnl", 0), "realised_pnl": bot_state.get("realised_pnl", 0), "total_return_pct": bot_state.get("return_pct", 0), "total_trades": bot_state.get("total_trades", 0), "wins": bot_state.get("wins", 0), "losses": bot_state.get("losses", 0), "win_rate": bot_state.get("win_rate", 0), "last_updated": "", "positions": data.get("positions", [])})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/watchlist")
def watchlist():
    raw = bot_state.get("watchlist", [])
    wlist = []
    if isinstance(raw, dict):
        for sym, w in raw.items():
            wlist.append({"symbol": sym.replace(".NS","").replace("-USD",""), "full_symbol": sym, "signal": w.get("signal","HOLD"), "price": w.get("price",0), "rsi": w.get("rsi",50)})
    elif isinstance(raw, list):
        wlist = raw
    return jsonify({"watchlist": wlist, "scanned_at": bot_state.get("scanned_at","")})

@app.route("/alerts")
def alerts():
    return jsonify({"alerts": bot_state.get("alerts", [])})

@app.route("/status")
def status():
    return jsonify({"status": "running", "connected": True})

@app.route("/ping")
def ping():
    return "pong", 200

@app.route("/trades")
def trades():
    t = bot_state.get("trades", bot_state.get("trade_log", []))
    return jsonify({"trades": t})

@app.route("/buy", methods=["POST"])
def buy():
    return jsonify({"status": "ok", "message": "Manual buy not supported in paper mode"})

@app.route("/sell", methods=["POST"])
def sell():
    return jsonify({"status": "ok", "message": "Manual sell not supported in paper mode"})

@app.route("/api/upload_state", methods=["POST"])
def upload_state():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data"}), 400
        for key in ["capital", "positions", "trades"]:
            if key in data:
                bot_state[key] = data[key]
        if "trades" in data:
            bot_state["trade_log"] = data["trades"]
        try:
            from db_state import save_state as db_save
            db_save(dict(bot_state))
        except Exception as e:
            print(f"[Upload] DB error: {e}")
        return jsonify({"status":"ok","trades":len(data.get("trades",[])),"positions":len(data.get("positions",{}))})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/download/trades")
def download_trades():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    try:
        trades = bot_state.get("trades", bot_state.get("trade_log", []))
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade History"
        if trades:
            headers = list(trades[0].keys())
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1a2135")
                cell.alignment = Alignment(horizontal="center")
            for t in trades:
                ws.append([t.get(h,"") for h in headers])
        else:
            ws.append(["No trades recorded yet"])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len+4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name="RSI_Bot_Trades.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/')
def index():
    return render_template_string(DASHBOARD_HTML)

def push_updates():
    while True:
        try:
            from db_state import load_trades as _lt
            all_trades = _lt()
            bot_state["trades"] = all_trades
            closed = [t for t in all_trades if t.get("action") == "SELL"]
            wins = [t for t in closed if (t.get("pnl") or 0) > 0]
            losses = [t for t in closed if (t.get("pnl") or 0) <= 0]
            bot_state["total_trades"] = len(closed)
            bot_state["wins"] = len(wins)
            bot_state["losses"] = len(losses)
            bot_state["win_rate"] = round(len(wins)/len(closed)*100,1) if closed else 0
        except Exception as e:
            print(f"[push] {e}")
        from db_state import load_state as _fresh
        _fs = _fresh()
        if _fs:
            bot_state['capital'] = _fs.get('capital', bot_state['capital'])
            bot_state['positions'] = _fs.get('positions', bot_state['positions'])
        socketio.emit('state_update', bot_state)
        time.sleep(5)

def start_dashboard():
    import logging
    logging.getLogger('werkzeug').setLevel(logging.ERROR)
    logging.getLogger('socketio').setLevel(logging.ERROR)
    logging.getLogger('engineio').setLevel(logging.ERROR)
    threading.Thread(target=push_updates, daemon=True).start()
    socketio.run(app, host='0.0.0.0', port=5000, log_output=False, debug=False, allow_unsafe_werkzeug=True)

@app.route("/upstox/callback")
def upstox_callback():
    code = request.args.get("code")
    if not code:
        return "No code received", 400
    from upstox_integration import get_access_token
    token = get_access_token(code)
    if token:
        return "<h2>OK Upstox Connected!</h2>"
    return "<h2>X Auth failed</h2>", 500

@app.route("/upstox/postback", methods=["POST"])
def upstox_postback():
    data = request.get_json(silent=True) or {}
    try:
        from upstox_integration import load_token, sync_to_bot
        token = load_token()
        if token:
            sync_to_bot(token)
    except Exception as e:
        print(f"[Upstox] {e}")
    return jsonify({"status": "ok"})

@app.route('/api/state')
def api_state():
    from db_state import load_state as _db
    s = _db() or {}
    return jsonify({'capital': s.get('capital', 100000), 'positions': len(s.get('positions', {}))})

if __name__ == "__main__":
    start_dashboard()




